#!/usr/bin/env python3
"""
Export OCR Review File v2 - Enhanced for Offline Reviewers

Includes:
- Source provenance (document title, collection, page numbers)
- Text snippets with context (paragraph around mention)
- Space for entity descriptions
- Grouped by cluster for easy batch review

Output:
- review_clusters.csv: One row per cluster with summary + reviewer columns
- review_details.csv: All variants with full context for each cluster
- Optional: XLSX with multiple sheets

Usage:
    python scripts/export_ocr_review_file_v2.py --output-dir review_export/ --limit 50
"""

import argparse
import csv
import hashlib
import os
import re
import sys
from datetime import datetime
from typing import List, Dict, Optional, Tuple

import psycopg2
from psycopg2.extras import RealDictCursor

sys.path.insert(0, '.')


# =============================================================================
# CONFIGURATION
# =============================================================================

CONTEXT_CHARS = 150  # Characters before/after mention for snippet
MAX_EXAMPLES_PER_CLUSTER = 10  # Max variant examples to show


def get_conn():
    return psycopg2.connect(
        host=os.environ.get('POSTGRES_HOST', 'localhost'),
        port=int(os.environ.get('POSTGRES_PORT', '5432')),
        dbname=os.environ.get('POSTGRES_DB', 'neh'),
        user=os.environ.get('POSTGRES_USER', 'neh'),
        password=os.environ.get('POSTGRES_PASSWORD', 'neh')
    )


def clean_text_for_csv(text: str) -> str:
    """Clean text for CSV export (remove newlines, normalize whitespace)."""
    if not text:
        return ''
    # Replace newlines with spaces
    text = re.sub(r'[\r\n]+', ' ', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_snippet(text: str, start: int, end: int, context_chars: int = CONTEXT_CHARS) -> str:
    """Extract a snippet around the mention with context."""
    if not text:
        return ''
    
    # Expand to context window
    snippet_start = max(0, start - context_chars)
    snippet_end = min(len(text), end + context_chars)
    
    snippet = text[snippet_start:snippet_end]
    
    # Add ellipsis if truncated
    if snippet_start > 0:
        snippet = '...' + snippet
    if snippet_end < len(text):
        snippet = snippet + '...'
    
    # Mark the mention
    mention_in_snippet_start = start - snippet_start + (3 if snippet_start > 0 else 0)
    mention_in_snippet_end = mention_in_snippet_start + (end - start)
    
    # Insert markers
    marked = (
        snippet[:mention_in_snippet_start] + 
        '>>>' + 
        snippet[mention_in_snippet_start:mention_in_snippet_end] + 
        '<<<' + 
        snippet[mention_in_snippet_end:]
    )
    
    return clean_text_for_csv(marked)


def get_cluster_details(conn, cluster_id: str) -> Dict:
    """Get detailed information about a cluster and its variants."""
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Get cluster info
    cur.execute("""
        SELECT 
            c.*,
            e.canonical_name as entity_name,
            e.entity_type,
            e.notes as entity_notes
        FROM ocr_variant_clusters c
        LEFT JOIN entities e ON e.id = c.canonical_entity_id
        WHERE c.cluster_id = %s
    """, (cluster_id,))
    
    cluster = dict(cur.fetchone())
    
    # Get variants with full context
    cur.execute("""
        SELECT 
            v.variant_key,
            v.raw_examples,
            v.mention_count,
            v.doc_count,
            v.avg_quality_score,
            v.current_entity_id,
            v.example_citations
        FROM ocr_cluster_variants v
        WHERE v.cluster_id = %s
        ORDER BY v.mention_count DESC
    """, (cluster_id,))
    
    variants = [dict(row) for row in cur.fetchall()]
    
    # Get example mentions with full context from mention_review_queue
    cur.execute("""
        SELECT 
            mrq.surface_norm,
            mrq.surface,
            mrq.chunk_id,
            mrq.document_id,
            mrq.start_char,
            mrq.end_char,
            mrq.context_excerpt,
            COALESCE(d.source_name, d.volume_key, 'Doc ' || d.id::text) as doc_title,
            col.title as collection_name,
            d.metadata->>'page' as page_number,
            c.text as chunk_text
        FROM mention_review_queue mrq
        JOIN documents d ON d.id = mrq.document_id
        JOIN collections col ON col.id = d.collection_id
        LEFT JOIN chunks c ON c.id = mrq.chunk_id
        WHERE mrq.surface_norm IN (
            SELECT unnest(raw_examples) FROM ocr_cluster_variants WHERE cluster_id = %s
        )
        ORDER BY mrq.document_id, mrq.start_char
        LIMIT %s
    """, (cluster_id, MAX_EXAMPLES_PER_CLUSTER))
    
    examples = [dict(row) for row in cur.fetchall()]
    
    # Enhance examples with snippets
    for ex in examples:
        if ex.get('chunk_text') and ex.get('start_char') is not None:
            ex['snippet'] = extract_snippet(
                ex['chunk_text'], 
                ex['start_char'], 
                ex['end_char'],
                CONTEXT_CHARS
            )
        else:
            ex['snippet'] = ex.get('context_excerpt', '')
    
    cluster['variants'] = variants
    cluster['examples'] = examples
    
    return cluster


def format_document_ref(doc_title: str, collection: str, page: Optional[int]) -> str:
    """Format a document reference string."""
    parts = []
    if collection:
        parts.append(f"[{collection}]")
    if doc_title:
        # Truncate long titles
        title = doc_title[:50] + '...' if len(doc_title) > 50 else doc_title
        parts.append(title)
    if page:
        parts.append(f"p.{page}")
    return ' '.join(parts) if parts else '(unknown)'


def export_enhanced_csv(conn, output_dir: str, limit: Optional[int] = None) -> Tuple[int, str, str]:
    """
    Export enhanced CSV files for offline review.
    
    Returns: (cluster_count, clusters_path, details_path)
    """
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Get pending clusters
    limit_clause = f"LIMIT {limit}" if limit else ""
    cur.execute(f"""
        SELECT cluster_id
        FROM ocr_variant_clusters
        WHERE status = 'pending'
        ORDER BY priority_score DESC
        {limit_clause}
    """)
    
    cluster_ids = [row['cluster_id'] for row in cur.fetchall()]
    
    if not cluster_ids:
        print("No pending clusters to export.")
        return 0, '', ''
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Prepare files
    clusters_path = os.path.join(output_dir, f'review_clusters_{timestamp}.csv')
    details_path = os.path.join(output_dir, f'review_details_{timestamp}.csv')
    
    # Cluster summary file (one row per cluster)
    cluster_rows = []
    detail_rows = []
    
    for cluster_id in cluster_ids:
        cluster = get_cluster_details(conn, cluster_id)
        
        # Build variant summary
        variant_summary = []
        for v in cluster['variants'][:5]:
            examples = v.get('raw_examples', [])
            if examples:
                variant_summary.append(f"{examples[0]} ({v['mention_count']} mentions)")
        
        # Build document summary
        doc_refs = set()
        for ex in cluster['examples']:
            ref = format_document_ref(
                ex.get('doc_title'),
                ex.get('collection_name'),
                ex.get('page_number')
            )
            doc_refs.add(ref)
        
        # Cluster row
        cluster_rows.append({
            'cluster_id': cluster_id,
            'proposed_name': cluster.get('proposed_canonical', ''),
            'variant_count': cluster.get('variant_count', 0),
            'total_mentions': cluster.get('total_mentions', 0),
            'doc_count': cluster.get('doc_count', 0),
            'recommendation': cluster.get('recommendation', ''),
            'priority_score': float(cluster.get('priority_score', 0)),
            'existing_entity_id': cluster.get('canonical_entity_id', ''),
            'existing_entity_name': cluster.get('entity_name', ''),
            'existing_entity_type': cluster.get('entity_type', ''),
            'variant_examples': ' | '.join(variant_summary),
            'document_sources': ' | '.join(list(doc_refs)[:5]),
            'danger_flags': 'MULTI_ENTITY' if cluster.get('maps_to_multiple_entities') else (
                'TYPE_CONFLICT' if cluster.get('has_type_conflict') else ''
            ),
            # Reviewer columns
            'review_decision': '',
            'entity_description': '',
            'reviewer_notes': '',
        })
        
        # Detail rows (one per example)
        for ex in cluster['examples']:
            detail_rows.append({
                'cluster_id': cluster_id,
                'proposed_name': cluster.get('proposed_canonical', ''),
                'variant_surface': ex.get('surface', ex.get('surface_norm', '')),
                'document_title': clean_text_for_csv(ex.get('doc_title', '')),
                'collection': ex.get('collection_name', ''),
                'page_number': ex.get('page_number', ''),
                'document_id': ex.get('document_id', ''),
                'chunk_id': ex.get('chunk_id', ''),
                'text_snippet': ex.get('snippet', ''),
                'char_position': f"{ex.get('start_char', '')}-{ex.get('end_char', '')}",
            })
        
        # If no examples from queue, add variant info as fallback
        if not cluster['examples']:
            for v in cluster['variants'][:3]:
                for raw in (v.get('raw_examples', []) or [])[:2]:
                    detail_rows.append({
                        'cluster_id': cluster_id,
                        'proposed_name': cluster.get('proposed_canonical', ''),
                        'variant_surface': raw,
                        'document_title': '(see variant citations)',
                        'collection': '',
                        'page_number': '',
                        'document_id': '',
                        'chunk_id': '',
                        'text_snippet': f"Variant '{raw}' - {v.get('mention_count', 0)} mentions in {v.get('doc_count', 0)} docs",
                        'char_position': '',
                    })
    
    # Write cluster summary
    with open(clusters_path, 'w', newline='', encoding='utf-8') as f:
        if cluster_rows:
            writer = csv.DictWriter(f, fieldnames=list(cluster_rows[0].keys()))
            writer.writeheader()
            writer.writerows(cluster_rows)
    
    # Write details
    with open(details_path, 'w', newline='', encoding='utf-8') as f:
        if detail_rows:
            writer = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
            writer.writeheader()
            writer.writerows(detail_rows)
    
    return len(cluster_ids), clusters_path, details_path


def export_xlsx(conn, output_dir: str, limit: Optional[int] = None) -> Tuple[int, str]:
    """
    Export to Excel with multiple sheets for easy review.
    
    Sheets:
    1. Clusters: Summary with reviewer columns
    2. Details: Full context for each variant
    3. Instructions: How to review
    """
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return 0, ''
    
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Get pending clusters
    limit_clause = f"LIMIT {limit}" if limit else ""
    cur.execute(f"""
        SELECT cluster_id
        FROM ocr_variant_clusters
        WHERE status = 'pending'
        ORDER BY priority_score DESC
        {limit_clause}
    """)
    
    cluster_ids = [row['cluster_id'] for row in cur.fetchall()]
    
    if not cluster_ids:
        print("No pending clusters to export.")
        return 0, ''
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    xlsx_path = os.path.join(output_dir, f'ocr_review_{timestamp}.xlsx')
    
    wb = openpyxl.Workbook()
    
    # =========================================================================
    # Sheet 1: Instructions
    # =========================================================================
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    
    instructions = [
        ("OCR Entity Review - Instructions", None),
        ("", None),
        ("PURPOSE:", None),
        ("Review OCR-extracted entity mentions and decide how to handle them.", None),
        ("", None),
        ("DECISION OPTIONS:", None),
        ("  APPROVE_MERGE", "Link these variants to the suggested existing entity"),
        ("  APPROVE_NEW_ENTITY", "Create a new entity for these variants"),
        ("  BLOCK", "These are junk/noise - block them from future extraction"),
        ("  (blank/DEFER)", "Skip for now - leave for later review"),
        ("", None),
        ("HOW TO REVIEW:", None),
        ("1. Go to 'Clusters' sheet - one row per group of similar variants", None),
        ("2. Look at 'variant_examples' and 'document_sources' columns", None),
        ("3. Check 'Details' sheet for full text snippets", None),
        ("4. Fill in 'review_decision' column", None),
        ("5. Optionally add 'entity_description' for new entities", None),
        ("6. Add any notes in 'reviewer_notes'", None),
        ("", None),
        ("TIPS:", None),
        ("- >>>text<<< markers show the mention in context", None),
        ("- Check if existing_entity_name matches the variants", None),
        ("- BLOCK generic terms like 'Center', 'Office', etc.", None),
        ("- Use entity_description for people: role, affiliation, dates", None),
        ("", None),
        (f"Export timestamp: {timestamp}", None),
        (f"Clusters to review: {len(cluster_ids)}", None),
    ]
    
    for row_idx, (text, note) in enumerate(instructions, 1):
        ws_instr.cell(row=row_idx, column=1, value=text)
        if note:
            ws_instr.cell(row=row_idx, column=2, value=note)
        if row_idx == 1:
            ws_instr.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        elif text.endswith(':'):
            ws_instr.cell(row=row_idx, column=1).font = Font(bold=True)
    
    ws_instr.column_dimensions['A'].width = 60
    ws_instr.column_dimensions['B'].width = 50
    
    # =========================================================================
    # Sheet 2: Clusters (main review sheet)
    # =========================================================================
    ws_clusters = wb.create_sheet("Clusters")
    
    headers = [
        'cluster_id', 'proposed_name', 'variant_count', 'total_mentions',
        'doc_count', 'recommendation', 'priority_score',
        'existing_entity_id', 'existing_entity_name', 'existing_entity_type',
        'variant_examples', 'document_sources', 'danger_flags',
        'review_decision', 'entity_description', 'reviewer_notes'
    ]
    
    # Header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_clusters.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    row_idx = 2
    for cluster_id in cluster_ids:
        cluster = get_cluster_details(conn, cluster_id)
        
        # Build summaries
        variant_summary = []
        for v in cluster['variants'][:5]:
            examples = v.get('raw_examples', [])
            if examples:
                variant_summary.append(f"{examples[0]} ({v['mention_count']})")
        
        doc_refs = set()
        for ex in cluster['examples']:
            ref = format_document_ref(
                ex.get('doc_title'),
                ex.get('collection_name'),
                ex.get('page_number')
            )
            doc_refs.add(ref)
        
        danger = ''
        if cluster.get('maps_to_multiple_entities'):
            danger = 'MULTI_ENTITY'
        elif cluster.get('has_type_conflict'):
            danger = 'TYPE_CONFLICT'
        
        row_data = [
            cluster_id,
            cluster.get('proposed_canonical', ''),
            cluster.get('variant_count', 0),
            cluster.get('total_mentions', 0),
            cluster.get('doc_count', 0),
            cluster.get('recommendation', ''),
            float(cluster.get('priority_score', 0)),
            cluster.get('canonical_entity_id', ''),
            cluster.get('entity_name', ''),
            cluster.get('entity_type', ''),
            ' | '.join(variant_summary),
            ' | '.join(list(doc_refs)[:5]),
            danger,
            '',  # review_decision
            '',  # entity_description
            '',  # reviewer_notes
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            ws_clusters.cell(row=row_idx, column=col_idx, value=value)
        
        row_idx += 1
    
    # Add data validation for review_decision
    dv = DataValidation(
        type="list",
        formula1='"APPROVE_MERGE,APPROVE_NEW_ENTITY,BLOCK,DEFER"',
        allow_blank=True
    )
    dv.error = "Please select: APPROVE_MERGE, APPROVE_NEW_ENTITY, BLOCK, or leave blank"
    ws_clusters.add_data_validation(dv)
    dv.add(f'N2:N{row_idx}')  # Column N = review_decision
    
    # Column widths
    col_widths = [18, 30, 12, 14, 10, 14, 12, 16, 25, 15, 50, 50, 12, 18, 40, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws_clusters.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header
    ws_clusters.freeze_panes = 'A2'
    
    # =========================================================================
    # Sheet 3: Details (full context)
    # =========================================================================
    ws_details = wb.create_sheet("Details")
    
    detail_headers = [
        'cluster_id', 'proposed_name', 'variant_surface',
        'document_title', 'collection', 'page_number',
        'document_id', 'text_snippet'
    ]
    
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_idx = 2
    for cluster_id in cluster_ids:
        cluster = get_cluster_details(conn, cluster_id)
        
        for ex in cluster['examples']:
            row_data = [
                cluster_id,
                cluster.get('proposed_canonical', ''),
                ex.get('surface', ex.get('surface_norm', '')),
                clean_text_for_csv(ex.get('doc_title', '')),
                ex.get('collection_name', ''),
                ex.get('page_number', ''),
                ex.get('document_id', ''),
                ex.get('snippet', ''),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                ws_details.cell(row=row_idx, column=col_idx, value=value)
            
            row_idx += 1
        
        # Fallback for clusters without queue examples
        if not cluster['examples']:
            for v in cluster['variants'][:3]:
                for raw in (v.get('raw_examples', []) or [])[:2]:
                    row_data = [
                        cluster_id,
                        cluster.get('proposed_canonical', ''),
                        raw,
                        '(from variant data)',
                        '',
                        '',
                        '',
                        f"Variant '{raw}' - {v.get('mention_count', 0)} mentions",
                    ]
                    for col_idx, value in enumerate(row_data, 1):
                        ws_details.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1
    
    # Column widths
    detail_widths = [18, 30, 25, 40, 15, 10, 12, 80]
    for col_idx, width in enumerate(detail_widths, 1):
        ws_details.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws_details.freeze_panes = 'A2'
    
    # Save
    wb.save(xlsx_path)
    
    return len(cluster_ids), xlsx_path


def main():
    parser = argparse.ArgumentParser(description='Export enhanced OCR review file')
    parser.add_argument('--output-dir', required=True, help='Output directory')
    parser.add_argument('--limit', type=int, help='Limit number of clusters')
    parser.add_argument('--format', choices=['csv', 'xlsx', 'both'], default='both',
                       help='Output format')
    args = parser.parse_args()
    
    os.makedirs(args.output_dir, exist_ok=True)
    
    conn = get_conn()
    
    print("=== Export Enhanced OCR Review File ===")
    print(f"Output directory: {args.output_dir}")
    print(f"Format: {args.format}")
    print(f"Limit: {args.limit or 'none'}")
    print()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    exported = {}
    
    if args.format in ('csv', 'both'):
        count, clusters_path, details_path = export_enhanced_csv(conn, args.output_dir, args.limit)
        if count > 0:
            print(f"Exported {count} clusters to:")
            print(f"  {clusters_path} (summary for decisions)")
            print(f"  {details_path} (full context)")
            exported['clusters_csv'] = clusters_path
            exported['details_csv'] = details_path
    
    if args.format in ('xlsx', 'both'):
        count, xlsx_path = export_xlsx(conn, args.output_dir, args.limit)
        if count > 0:
            print(f"Exported {count} clusters to:")
            print(f"  {xlsx_path}")
            exported['xlsx'] = xlsx_path
    
    # Write manifest
    if exported:
        manifest_path = os.path.join(args.output_dir, f'manifest_{timestamp}.txt')
        with open(manifest_path, 'w') as f:
            f.write(f"Export timestamp: {timestamp}\n")
            f.write(f"Limit: {args.limit or 'none'}\n\n")
            f.write("Files:\n")
            for key, path in exported.items():
                file_hash = hashlib.md5(open(path, 'rb').read()).hexdigest()
                f.write(f"  {key}: {os.path.basename(path)}\n")
                f.write(f"    MD5: {file_hash}\n")
        
        print(f"\nManifest: {manifest_path}")
    
    print("\n" + "="*60)
    print("NEXT STEPS FOR REVIEWER:")
    print("="*60)
    print("1. Open the review file (XLSX recommended)")
    print("2. Read 'Instructions' sheet first")
    print("3. Review 'Clusters' sheet - fill in 'review_decision' column")
    print("4. Check 'Details' sheet for text context")
    print("5. Add 'entity_description' for new entities")
    print("6. Save and return for import")
    print()
    print("Run import with:")
    print(f"  python scripts/apply_ocr_adjudication.py <reviewed_file>")
    
    conn.close()


if __name__ == '__main__':
    main()
