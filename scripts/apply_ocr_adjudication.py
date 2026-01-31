#!/usr/bin/env python3
"""
Apply OCR Adjudication Decisions

Imports reviewed CSV/Excel file and applies decisions:
- APPROVE_MERGE: Add variant aliases to existing entity
- APPROVE_NEW_ENTITY: Create new entity with aliases
- BLOCK: Add to blocklist (never propose again)
- DEFER: No action (keep in pending state)

Usage:
    python scripts/apply_ocr_adjudication.py review_export/review_file_xxx.xlsx
    python scripts/apply_ocr_adjudication.py review_export/clusters_xxx.csv --dry-run
"""

import argparse
import csv
import hashlib
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import psycopg2
from psycopg2.extras import Json

sys.path.insert(0, '.')
from retrieval.surface_norm import normalize_surface


def get_conn():
    return psycopg2.connect(
        host='localhost', port=5432, dbname='neh', user='neh', password='neh'
    )


def get_or_create_ocr_source(conn) -> int:
    """
    Get or create a source ID for OCR adjudication.
    
    Uses the first available source_id from the database.
    In production, you might want to create a dedicated source.
    """
    cur = conn.cursor()
    
    # Try to get any existing source_id
    cur.execute("SELECT id FROM concordance_sources LIMIT 1")
    row = cur.fetchone()
    if row:
        return row[0]
    
    # Fallback: get from entities table
    cur.execute("SELECT DISTINCT source_id FROM entities WHERE source_id IS NOT NULL LIMIT 1")
    row = cur.fetchone()
    if row:
        return row[0]
    
    # Last resort: return 1 (may fail if doesn't exist)
    return 1


def compute_file_hash(filepath: str) -> str:
    """Compute MD5 hash of file."""
    with open(filepath, 'rb') as f:
        return hashlib.md5(f.read()).hexdigest()


def read_csv_decisions(filepath: str) -> List[Dict]:
    """Read decisions from CSV file."""
    decisions = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            decision = row.get('review_decision', '').strip().upper()
            if decision and decision != 'DEFER':
                decisions.append({
                    'cluster_id': row.get('cluster_id'),
                    'decision': decision,
                    'notes': row.get('reviewer_notes', ''),
                    'entity_description': row.get('entity_description', ''),
                    'proposed_canonical': row.get('proposed_canonical') or row.get('proposed_name'),
                    'canonical_entity_id': row.get('canonical_entity_id') or row.get('existing_entity_id'),
                    'entity_name': row.get('entity_name') or row.get('existing_entity_name')
                })
    return decisions


def read_xlsx_decisions(filepath: str) -> List[Dict]:
    """Read decisions from Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return []
    
    wb = openpyxl.load_workbook(filepath)
    
    # Try to find the Clusters sheet, or use active sheet
    if 'Clusters' in wb.sheetnames:
        ws = wb['Clusters']
    else:
        ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    decisions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        decision = str(row_dict.get('review_decision', '') or '').strip().upper()
        if decision and decision != 'DEFER':
            decisions.append({
                'cluster_id': row_dict.get('cluster_id'),
                'decision': decision,
                'notes': row_dict.get('reviewer_notes', ''),
                'entity_description': row_dict.get('entity_description', ''),
                'proposed_canonical': row_dict.get('proposed_canonical') or row_dict.get('proposed_name'),
                'canonical_entity_id': row_dict.get('canonical_entity_id') or row_dict.get('existing_entity_id'),
                'entity_name': row_dict.get('entity_name') or row_dict.get('existing_entity_name')
            })
    
    return decisions


def get_cluster_variants(conn, cluster_id: str) -> List[Dict]:
    """Get all variants for a cluster."""
    cur = conn.cursor()
    cur.execute("""
        SELECT variant_key, raw_examples, mention_count, current_entity_id
        FROM ocr_cluster_variants
        WHERE cluster_id = %s
    """, (cluster_id,))
    
    variants = []
    for row in cur.fetchall():
        variants.append({
            'variant_key': row[0],
            'raw_examples': row[1] or [],
            'mention_count': row[2],
            'current_entity_id': row[3]
        })
    return variants


def apply_approve_merge(conn, cluster_id: str, entity_id: int, notes: str, reviewer: str) -> Tuple[int, str]:
    """
    Apply APPROVE_MERGE decision.
    
    Adds variant aliases to the specified entity (does NOT create new entity).
    Returns (count of aliases added, status message).
    """
    cur = conn.cursor()
    
    # Get cluster variants
    variants = get_cluster_variants(conn, cluster_id)
    if not variants:
        return 0, "No variants found"
    
    # Verify entity exists and get source_id
    cur.execute("SELECT canonical_name, source_id FROM entities WHERE id = %s", (entity_id,))
    row = cur.fetchone()
    if not row:
        return 0, f"Entity {entity_id} not found"
    
    entity_name, source_id = row
    if not source_id:
        source_id = get_or_create_ocr_source(conn)
    
    added = 0
    
    for v in variants:
        # Add to allowlist (this is the key action - maps variant to entity)
        cur.execute("""
            INSERT INTO ocr_variant_allowlist (variant_key, entity_id, reason, source)
            VALUES (%s, %s, %s, 'adjudication')
            ON CONFLICT (variant_key, entity_id) DO NOTHING
        """, (v['variant_key'], entity_id, f"Merged from cluster {cluster_id}"))
        
        # Also try to add aliases (may fail due to schema constraints)
        for raw_example in v['raw_examples'][:3]:  # Limit to 3 examples
            alias_norm = normalize_surface(raw_example)
            if not alias_norm:
                continue
            
            try:
                cur.execute("""
                    INSERT INTO entity_aliases (
                        source_id, entity_id, alias, alias_norm, alias_type,
                        notes, is_auto_match, is_matchable
                    ) VALUES (%s, %s, %s, %s, 'ocr_variant', %s, TRUE, TRUE)
                    ON CONFLICT DO NOTHING
                """, (
                    source_id,
                    entity_id,
                    raw_example,
                    alias_norm,
                    f"From OCR cluster {cluster_id}: {notes}"
                ))
                if cur.rowcount > 0:
                    added += 1
            except Exception as e:
                # Alias creation may fail due to schema constraints
                # The allowlist entry is the key, aliases are bonus
                pass
    
    # Update cluster status
    cur.execute("""
        UPDATE ocr_variant_clusters
        SET status = 'approved', review_decision = 'APPROVE_MERGE',
            reviewed_by = %s, reviewed_at = NOW()
        WHERE cluster_id = %s
    """, (reviewer, cluster_id))
    
    # Record event
    cur.execute("""
        INSERT INTO ocr_review_events (
            event_type, cluster_id, entity_id, decision, reviewer, payload
        ) VALUES ('cluster_review', %s, %s, 'APPROVE_MERGE', %s, %s)
    """, (cluster_id, entity_id, reviewer, Json({
        'notes': notes,
        'variants_added': added,
        'entity_name': entity_name
    })))
    
    conn.commit()
    return added, f"Added {added} aliases to entity '{entity_name}'"


def apply_approve_new_entity(conn, cluster_id: str, proposed_name: str, notes: str, reviewer: str, entity_description: str = '') -> Tuple[int, str]:
    """
    Apply APPROVE_NEW_ENTITY decision.
    
    Creates a new entity with the proposed canonical name and adds variant aliases.
    Returns (entity_id, status message).
    
    Note: This function creates new entities. Due to schema constraints requiring
    source_id, it uses a default source. In production, you may want a dedicated
    "OCR adjudication" source.
    """
    cur = conn.cursor()
    
    # Get cluster variants
    variants = get_cluster_variants(conn, cluster_id)
    if not variants:
        return 0, "No variants found"
    
    # Get source_id for new entity
    source_id = get_or_create_ocr_source(conn)
    
    # Build entity notes combining cluster info and reviewer description
    entity_notes = f"Created from OCR cluster {cluster_id}"
    if entity_description:
        entity_notes = f"{entity_description}\n\n[{entity_notes}]"
    if notes:
        entity_notes = f"{entity_notes}\nReviewer notes: {notes}"
    
    # Create new entity
    try:
        cur.execute("""
            INSERT INTO entities (source_id, canonical_name, entity_type, notes)
            VALUES (%s, %s, 'other', %s)
            RETURNING id
        """, (source_id, proposed_name, entity_notes))
        entity_id = cur.fetchone()[0]
    except Exception as e:
        return 0, f"Failed to create entity: {e}"
    
    # Add variant allowlist entries (the key action)
    for v in variants:
        cur.execute("""
            INSERT INTO ocr_variant_allowlist (variant_key, entity_id, reason, source)
            VALUES (%s, %s, %s, 'adjudication')
            ON CONFLICT (variant_key, entity_id) DO NOTHING
        """, (v['variant_key'], entity_id, f"New entity from cluster {cluster_id}"))
    
    # Try to add aliases (may fail due to schema constraints, but allowlist is key)
    added = 0
    canonical_norm = normalize_surface(proposed_name)
    
    for v in variants:
        for raw_example in v['raw_examples'][:3]:
            alias_norm = normalize_surface(raw_example)
            if not alias_norm:
                continue
            
            try:
                cur.execute("""
                    INSERT INTO entity_aliases (
                        source_id, entity_id, alias, alias_norm, alias_type,
                        notes, is_auto_match, is_matchable
                    ) VALUES (%s, %s, %s, %s, 'ocr_variant', %s, TRUE, TRUE)
                    ON CONFLICT DO NOTHING
                """, (source_id, entity_id, raw_example, alias_norm, 
                      f"From OCR cluster {cluster_id}"))
                if cur.rowcount > 0:
                    added += 1
            except Exception:
                # Alias creation may fail due to schema constraints
                pass
    
    # Update cluster status
    cur.execute("""
        UPDATE ocr_variant_clusters
        SET status = 'approved', review_decision = 'APPROVE_NEW_ENTITY',
            canonical_entity_id = %s, reviewed_by = %s, reviewed_at = NOW()
        WHERE cluster_id = %s
    """, (entity_id, reviewer, cluster_id))
    
    # Record event
    cur.execute("""
        INSERT INTO ocr_review_events (
            event_type, cluster_id, entity_id, decision, reviewer, payload
        ) VALUES ('cluster_review', %s, %s, 'APPROVE_NEW_ENTITY', %s, %s)
    """, (cluster_id, entity_id, reviewer, Json({
        'notes': notes,
        'entity_name': proposed_name,
        'aliases_added': added
    })))
    
    conn.commit()
    return entity_id, f"Created entity {entity_id} '{proposed_name}' with {added} aliases"


def apply_block(conn, cluster_id: str, notes: str, reviewer: str) -> Tuple[int, str]:
    """
    Apply BLOCK decision.
    
    Adds all variants to blocklist.
    Returns (count blocked, status message).
    """
    cur = conn.cursor()
    
    # Get cluster variants
    variants = get_cluster_variants(conn, cluster_id)
    if not variants:
        return 0, "No variants found"
    
    blocked = 0
    for v in variants:
        cur.execute("""
            INSERT INTO ocr_variant_blocklist (variant_key, block_type, cluster_id, reason, source)
            VALUES (%s, 'cluster', %s, %s, 'adjudication')
            ON CONFLICT DO NOTHING
        """, (v['variant_key'], cluster_id, notes or f"Blocked cluster {cluster_id}"))
        if cur.rowcount > 0:
            blocked += 1
    
    # Update cluster status
    cur.execute("""
        UPDATE ocr_variant_clusters
        SET status = 'blocked', review_decision = 'BLOCK',
            reviewed_by = %s, reviewed_at = NOW()
        WHERE cluster_id = %s
    """, (reviewer, cluster_id))
    
    # Record event
    cur.execute("""
        INSERT INTO ocr_review_events (
            event_type, cluster_id, decision, reviewer, payload
        ) VALUES ('cluster_review', %s, 'BLOCK', %s, %s)
    """, (cluster_id, reviewer, Json({
        'notes': notes,
        'variants_blocked': blocked
    })))
    
    conn.commit()
    return blocked, f"Blocked {blocked} variants"


def main():
    parser = argparse.ArgumentParser(description='Apply OCR adjudication decisions')
    parser.add_argument('input_file', help='Reviewed CSV or Excel file')
    parser.add_argument('--reviewer', default='adjudication_import', help='Reviewer name')
    parser.add_argument('--dry-run', action='store_true', help='Don\'t apply, just report')
    args = parser.parse_args()
    
    if not os.path.exists(args.input_file):
        print(f"ERROR: File not found: {args.input_file}")
        sys.exit(1)
    
    # Read decisions
    ext = os.path.splitext(args.input_file)[1].lower()
    if ext == '.xlsx':
        decisions = read_xlsx_decisions(args.input_file)
    elif ext == '.csv':
        decisions = read_csv_decisions(args.input_file)
    else:
        print(f"ERROR: Unsupported file format: {ext}")
        sys.exit(1)
    
    print("=== Apply OCR Adjudication ===")
    print(f"Input file: {args.input_file}")
    print(f"File hash: {compute_file_hash(args.input_file)}")
    print(f"Reviewer: {args.reviewer}")
    print(f"Dry run: {args.dry_run}")
    print()
    print(f"Found {len(decisions)} decisions to apply")
    
    if not decisions:
        print("No decisions to apply.")
        return
    
    conn = get_conn()
    
    # Group by decision type
    by_type = {}
    for d in decisions:
        dtype = d['decision']
        if dtype not in by_type:
            by_type[dtype] = []
        by_type[dtype].append(d)
    
    print("\nBy decision type:")
    for dtype, items in by_type.items():
        print(f"  {dtype}: {len(items)}")
    
    if args.dry_run:
        print("\n[DRY RUN] Would apply the following:")
        for d in decisions[:10]:
            print(f"  {d['cluster_id']}: {d['decision']}")
        if len(decisions) > 10:
            print(f"  ... and {len(decisions) - 10} more")
        return
    
    # Apply decisions
    print("\nApplying decisions...")
    
    results = {'success': 0, 'error': 0}
    
    for d in decisions:
        cluster_id = d['cluster_id']
        decision = d['decision']
        notes = d.get('notes', '')
        
        try:
            if decision == 'APPROVE_MERGE':
                entity_id = d.get('canonical_entity_id')
                if not entity_id:
                    print(f"  SKIP {cluster_id}: No entity_id for APPROVE_MERGE")
                    results['error'] += 1
                    continue
                
                count, msg = apply_approve_merge(conn, cluster_id, int(entity_id), notes, args.reviewer)
                print(f"  [OK] {cluster_id}: {msg}")
                results['success'] += 1
                
            elif decision == 'APPROVE_NEW_ENTITY':
                proposed = d.get('proposed_canonical') or d.get('entity_name')
                if not proposed:
                    print(f"  SKIP {cluster_id}: No proposed name for APPROVE_NEW_ENTITY")
                    results['error'] += 1
                    continue
                
                entity_description = d.get('entity_description', '')
                entity_id, msg = apply_approve_new_entity(conn, cluster_id, proposed, notes, args.reviewer, entity_description)
                print(f"  [OK] {cluster_id}: {msg}")
                results['success'] += 1
                
            elif decision == 'BLOCK':
                count, msg = apply_block(conn, cluster_id, notes, args.reviewer)
                print(f"  [OK] {cluster_id}: {msg}")
                results['success'] += 1
                
            else:
                print(f"  SKIP {cluster_id}: Unknown decision '{decision}'")
                results['error'] += 1
                
        except Exception as e:
            print(f"  [ERROR] {cluster_id}: Error - {e}")
            results['error'] += 1
            conn.rollback()
    
    print()
    print("=== RESULTS ===")
    print(f"  Success: {results['success']}")
    print(f"  Errors: {results['error']}")
    
    # Record the import event
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO ocr_review_events (
            event_type, decision, reviewer, source_file, source_file_hash, payload
        ) VALUES ('import', 'BATCH_IMPORT', %s, %s, %s, %s)
    """, (
        args.reviewer,
        os.path.basename(args.input_file),
        compute_file_hash(args.input_file),
        Json({
            'total_decisions': len(decisions),
            'success': results['success'],
            'errors': results['error']
        })
    ))
    conn.commit()
    
    conn.close()


if __name__ == '__main__':
    main()
